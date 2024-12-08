from pathlib import Path
import itertools
import typing
import rich
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def remove_empty_first_rows(worksheet: Worksheet):
    row_idx = 1
    # Loop until the first non-empty row is found
    while row_idx <= worksheet.max_row:
        # Check if all cells in the row are empty
        row = worksheet[row_idx]
        if all(cell.value is None for cell in row):
            # If the row is empty, delete it
            worksheet.delete_rows(row_idx)
        else:
            # If a non-empty row is found, stop deleting
            break

    return worksheet


def load_workbook(path: Path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    return workbook


def worksheets(workbook: openpyxl.Workbook):
    for sheet in workbook.sheetnames:
        yield workbook[sheet]


class SchemaInfo(typing.TypedDict):
    column: int
    overall: typing.NotRequired[typing.Optional[int]]


class SubjectSchema(typing.TypedDict):
    mid_term_score: SchemaInfo
    exam_score: SchemaInfo
    total_score: SchemaInfo


class BroadSheetSchema(typing.TypedDict):
    term: str
    subjects: typing.Dict[str, SubjectSchema]
    aggregates: typing.Dict[str, SchemaInfo]
    teachers_comment: SchemaInfo
    coordinators_comment: SchemaInfo


EXTERNAL_TO_INTERNAL_MAPPING = {
    "mid": "mid_term_score",
    "exam": "exam_score",
    "total": "total_score",
    "mid %": "mid term %",
    "mid total": "mid term total",
    "sim %": "sum total %",
    "sum %": "sum total %",
    "1st term": "1st term total",
    "2nd term": "2nd term total",
    "3rd term": "3rd term total",
    "cumtotal": "cumulative (session) total",
    "av. total": "average total",
    "av. %": "average %"
}


def _to_internal(val: str) -> str:
    val = val.strip().lower()
    return EXTERNAL_TO_INTERNAL_MAPPING.get(val, val)


def get_broadsheet_schema(worksheet: Worksheet):
    # Use typed dict for detailed typing and dictionary data access
    schema = BroadSheetSchema(
        term=worksheet.title.strip().title(),
        subjects={},
        aggregates={},
        teachers_comment={},
        coordinators_comment={},
    )
    last_column_index = None

    # Split the columns into batches of 3, because most columns (in row 2) have 3 sub columns (in row 3).
    # We limit the scope we need to iterate over to (row2, column3) to (row3, col*), as that is
    # the cell range in which the schema data we need to extract lies.
    # Simply put, just the 2nd and 3rd row are what we need to extract the column schema schema.
    for cols in itertools.batched(
        worksheet.iter_cols(min_col=3, min_row=1, max_row=3), n=3
    ):
        previous_title = None
        for col in cols:
            title = col[0].value
            sub_title = col[1].value
            sub_title_column_index = col[1].column
            overall = col[2].value

            if title:
                title = _to_internal(title)
                if title not in schema["subjects"]:
                    schema["subjects"][title] = {}
            else:
                if not previous_title:
                    if not sub_title:
                        continue

                    sub_title = _to_internal(sub_title)
                    if "comment" in sub_title:
                        continue

                    if sub_title not in schema["aggregates"]:
                        schema["aggregates"][sub_title] = {}

                    schema["aggregates"][sub_title]["column"] = sub_title_column_index
                    schema["aggregates"][sub_title]["overall"] = overall
                    last_column_index = sub_title_column_index
                    continue
                else:
                    title = previous_title

            if not sub_title:
                continue

            sub_title = _to_internal(sub_title)
            schema["subjects"][title][sub_title] = {"column": sub_title_column_index}
            schema["subjects"][title][sub_title]["overall"] = overall
            last_column_index = sub_title_column_index
            previous_title = title

    if last_column_index is not None:
        # Comments are anticipated to always be in the last column
        schema["teachers_comment"]["column"] = last_column_index + 1
        schema["coordinators_comment"]["column"] = last_column_index + 2
    return schema


class StudentInfo(typing.TypedDict):
    name: str
    row: int


def students(worksheet: Worksheet):
    for row in worksheet.iter_rows(min_row=4, min_col=2, max_col=2):
        name = row[0].value
        if not name:
            continue
        yield StudentInfo(
            name=name.strip().title(),
            row=row[0].row,
        )


Value = typing.Optional[typing.Union[int, float]]


class SubjectScore(typing.TypedDict):
    mid_term_score: Value
    exam_score: Value
    total_score: Value


SubjectsScores = typing.Dict[str, SubjectScore]
AggregatesValues = typing.Dict[str, Value]


class StudentResult(typing.TypedDict):
    term: str
    student: str
    subjects: SubjectsScores
    aggregates: AggregatesValues
    teachers_comment: typing.Optional[str]
    coordinators_comment: typing.Optional[str]


def get_subjects_scores_for_student(
    worksheet: Worksheet,
    student_row_index: int,
    subjects_schemas: typing.Dict[str, SubjectSchema],
):
    subjects_scores: SubjectsScores = {}
    for subject, subject_schema in subjects_schemas.items():
        mid_term_score_column_index = subject_schema["mid_term_score"]["column"]
        exam_score_column_index = subject_schema["exam_score"]["column"]
        total_score_column_index = subject_schema["total_score"]["column"]
        mid_term_score = worksheet.cell(
            student_row_index, mid_term_score_column_index
        ).value
        exam_score = worksheet.cell(student_row_index, exam_score_column_index).value
        total_score = worksheet.cell(student_row_index, total_score_column_index).value
        subject_score = SubjectScore(
            mid_term_score=mid_term_score,
            exam_score=exam_score,
            total_score=total_score,
        )
        subjects_scores[subject] = subject_score
    return subjects_scores


def get_aggregates_values(
    worksheet: Worksheet,
    student_row_index: int,
    aggregates_schemas: typing.Dict[str, SchemaInfo],
):
    aggregates_values: AggregatesValues = {}
    for aggregate, aggregate_schema in aggregates_schemas.items():
        aggregate_column_index = aggregate_schema["column"]
        aggregate_value = worksheet.cell(
            student_row_index, aggregate_column_index
        ).value
        aggregates_values[aggregate] = aggregate_value
    return aggregates_values


def get_comment_value(
    worksheet: Worksheet, student_row_index: int, comment_column_index: int
):
    comment = worksheet.cell(student_row_index, comment_column_index).value
    if not comment:
        return None
    return comment.strip()


def student_results(
    worksheet: Worksheet, sheet_schema: typing.Optional[BroadSheetSchema] = None
):
    broadsheet_schema = sheet_schema or get_broadsheet_schema(worksheet)
    for student in students(worksheet):
        student_row_index = student["row"]
        student_name = student["name"]
        subjects_schemas = broadsheet_schema["subjects"]
        aggregates_schemas = broadsheet_schema["aggregates"]
        teachers_comment_schema = broadsheet_schema["teachers_comment"]
        coordinators_comment_schema = broadsheet_schema["coordinators_comment"]

        subjects_scores = get_subjects_scores_for_student(
            worksheet=worksheet,
            student_row_index=student_row_index,
            subjects_schemas=subjects_schemas,
        )
        aggregates_values = get_aggregates_values(
            worksheet=worksheet,
            student_row_index=student_row_index,
            aggregates_schemas=aggregates_schemas,
        )
        teachers_comment = get_comment_value(
            worksheet=worksheet,
            student_row_index=student_row_index,
            comment_column_index=teachers_comment_schema["column"],
        )
        coordinators_comment = get_comment_value(
            worksheet=worksheet,
            student_row_index=student_row_index,
            comment_column_index=coordinators_comment_schema["column"],
        )
        result = StudentResult(
            term=broadsheet_schema["term"],
            student=student_name,
            subjects=subjects_scores,
            aggregates=aggregates_values,
            teachers_comment=teachers_comment,
            coordinators_comment=coordinators_comment,
        )
        yield result


def extract_broadsheet_data(broadsheet: str):
    workbook = load_workbook(Path(broadsheet).resolve())
    sheets_data = {}
    for worksheet in worksheets(workbook):
        worksheet = remove_empty_first_rows(worksheet)
        broadsheet_schema = get_broadsheet_schema(worksheet)
        results: typing.List[StudentResult] = []
        for student_result in student_results(
            worksheet, sheet_schema=broadsheet_schema
        ):  
            results.append(student_result)

        sheets_data[broadsheet_schema["term"]] = results
    return sheets_data
